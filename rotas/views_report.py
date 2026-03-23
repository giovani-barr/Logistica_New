"""
Views para o Editor de Layout de Relatórios (Report Designer)
"""
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST, require_http_methods
from django.views.decorators.csrf import ensure_csrf_cookie
from django.contrib import messages
from rotas.models import LayoutReport, ReportHistory, QueryFirebird, Rota, PedidoFirebird
from rotas.report_renderer import ReportRenderer, generate_pdf, generate_preview_pdf
import json
import time
from datetime import datetime


@login_required
def report_designer_list(request):
    """Lista todos os layouts de relatório do usuário"""
    layouts = LayoutReport.objects.filter(usuario=request.user, ativo=True).order_by('-data_modificacao')
    
    context = {
        'layouts': layouts,
        'total_layouts': layouts.count()
    }
    return render(request, 'rotas/report_designer_list.html', context)


@login_required
def report_designer_editor(request, pk=None):
    """Editor visual de layout de relatório"""
    layout = None
    is_new = pk is None
    
    if pk:
        layout = get_object_or_404(LayoutReport, pk=pk, usuario=request.user)
    
    # Buscar queries disponíveis para associar ao layout
    queries = QueryFirebird.objects.filter(usuario=request.user, ativo=True).order_by('-data_modificacao')
    
    # Buscar campos do SQL ativo (mesmo SQL usado nos cards)
    from rotas.views import _extract_sql_aliases, SQL_CAMPOS_DISPONIVEIS
    
    query_ativa = QueryFirebird.objects.filter(usuario=request.user, ativo=True).order_by('-data_modificacao').first()
    
    campos_disponiveis = []
    sql_ativo = ''
    
    if query_ativa:
        sql_ativo = query_ativa.sql
        campos_disponiveis = _extract_sql_aliases(sql_ativo)
    
    # Se não encontrou campos no SQL ou não há SQL, usar campos padrão
    if not campos_disponiveis:
        campos_disponiveis = SQL_CAMPOS_DISPONIVEIS
    
    context = {
        'layout': layout,
        'is_new': is_new,
        'queries': queries,
        'query_ativa': query_ativa,
        'campos_disponiveis': json.dumps(campos_disponiveis, ensure_ascii=False),
        'campos_disponiveis_list': campos_disponiveis,  # Lista Python para template
        'campos_count': len(campos_disponiveis),
        'layout_json': json.dumps(layout.get_layout_structure() if layout else {}, ensure_ascii=False) if layout else '{}',
        'page_sizes': LayoutReport.PAGE_SIZE_CHOICES,
        'orientations': LayoutReport.ORIENTATION_CHOICES,
    }
    return render(request, 'rotas/report_designer_editor.html', context)


@login_required
@require_POST
def report_designer_save(request):
    """Salvar layout de relatório (create/update)"""
    try:
        data = json.loads(request.body.decode('utf-8'))
    except (json.JSONDecodeError, UnicodeDecodeError):
        return JsonResponse({'success': False, 'message': 'Dados inválidos'}, status=400)
    
    layout_id = data.get('id')
    nome = data.get('nome', '').strip()
    descricao = data.get('descricao', '').strip()
    query_id = data.get('query_id')
    layout_json = data.get('layout_json', {})
    page_size = data.get('page_size', 'A4')
    orientation = data.get('orientation', 'portrait')
    padrao = data.get('padrao', False)
    
    # Validações
    if not nome:
        return JsonResponse({'success': False, 'message': 'Nome é obrigatório'}, status=400)
    
    # Validar margens
    try:
        margin_top = float(data.get('margin_top', 20.0))
        margin_bottom = float(data.get('margin_bottom', 20.0))
        margin_left = float(data.get('margin_left', 15.0))
        margin_right = float(data.get('margin_right', 15.0))
        default_font_size = int(data.get('default_font_size', 10))
    except (ValueError, TypeError):
        return JsonResponse({'success': False, 'message': 'Margens e tamanho de fonte devem ser números válidos'}, status=400)
    
    # Buscar query
    query = None
    if query_id:
        try:
            query = QueryFirebird.objects.get(pk=query_id, usuario=request.user)
        except QueryFirebird.DoesNotExist:
            pass
    
    # Criar ou atualizar layout
    if layout_id:
        # Atualizar existente
        try:
            layout = LayoutReport.objects.get(pk=layout_id, usuario=request.user)
            layout.nome = nome
            layout.descricao = descricao
            layout.query = query
            layout.layout_json = layout_json
            layout.page_size = page_size
            layout.orientation = orientation
            layout.margin_top = margin_top
            layout.margin_bottom = margin_bottom
            layout.margin_left = margin_left
            layout.margin_right = margin_right
            layout.default_font_family = data.get('default_font_family', 'Helvetica')
            layout.default_font_size = default_font_size
            layout.padrao = padrao
            layout.versao += 1
            layout.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Layout "{nome}" atualizado com sucesso!',
                'layout_id': layout.id,
                'versao': layout.versao
            })
        except LayoutReport.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Layout não encontrado'}, status=404)
    else:
        # Criar novo
        layout = LayoutReport.objects.create(
            usuario=request.user,
            nome=nome,
            descricao=descricao,
            query=query,
            layout_json=layout_json,
            page_size=page_size,
            orientation=orientation,
            margin_top=margin_top,
            margin_bottom=margin_bottom,
            margin_left=margin_left,
            margin_right=margin_right,
            default_font_family=data.get('default_font_family', 'Helvetica'),
            default_font_size=default_font_size,
            padrao=padrao
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Layout "{nome}" criado com sucesso!',
            'layout_id': layout.id,
            'versao': layout.versao
        })


@login_required
@require_POST
def report_designer_delete(request, pk):
    """Deletar layout de relatório (soft delete)"""
    layout = get_object_or_404(LayoutReport, pk=pk, usuario=request.user)
    nome_layout = layout.nome
    layout.ativo = False
    layout.save()
    
    messages.success(request, f'Layout "{nome_layout}" deletado com sucesso!')
    return redirect('rotas:report_designer_list')


@login_required
def report_designer_duplicate(request, pk):
    """Duplicar um layout existente"""
    layout_original = get_object_or_404(LayoutReport, pk=pk, usuario=request.user)
    
    # Criar cópia
    layout_novo = LayoutReport.objects.create(
        usuario=request.user,
        nome=f"{layout_original.nome} (Cópia)",
        descricao=layout_original.descricao,
        query=layout_original.query,
        layout_json=layout_original.layout_json.copy() if layout_original.layout_json else {},
        page_size=layout_original.page_size,
        orientation=layout_original.orientation,
        margin_top=layout_original.margin_top,
        margin_bottom=layout_original.margin_bottom,
        margin_left=layout_original.margin_left,
        margin_right=layout_original.margin_right,
        default_font_family=layout_original.default_font_family,
        default_font_size=layout_original.default_font_size,
        padrao=False
    )
    
    messages.success(request, f'Layout "{layout_novo.nome}" criado como cópia!')
    return redirect('rotas:report_designer_editor', pk=layout_novo.pk)


@login_required
@require_http_methods(["GET"])
def report_designer_load_fields(request):
    """Carregar campos disponíveis de uma query específica"""
    query_id = request.GET.get('query_id')
    
    if not query_id:
        return JsonResponse({'success': False, 'message': 'Query ID não fornecido'}, status=400)
    
    try:
        query = QueryFirebird.objects.get(pk=query_id, usuario=request.user)
    except QueryFirebird.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Query não encontrada'}, status=404)
    
    # Extrair campos do SQL
    from rotas.views import _extract_sql_aliases
    campos = _extract_sql_aliases(query.sql)
    
    if not campos:
        # Campos padrão
        campos = [
            'Pedido', 'Cliente nome', 'Endereco entrega', 'Fone cliente',
            'LATITUDE', 'LONGITUDE', 'Observacao cliente', 'Nome vendedor'
        ]
    
    return JsonResponse({
        'success': True,
        'campos': campos,
        'query_nome': query.nome_query
    })


@login_required
@require_POST
def report_designer_preview(request):
    """Gerar preview do relatório em PDF usando ReportLab"""
    try:
        data = json.loads(request.body.decode('utf-8'))
    except (json.JSONDecodeError, UnicodeDecodeError):
        return JsonResponse({'success': False, 'message': 'Dados inválidos'}, status=400)
    
    layout_id = data.get('layout_id')
    layout_json = data.get('layout_json', {})
    
    if not layout_id:
        return JsonResponse({'success': False, 'message': 'Layout ID não fornecido'}, status=400)
    
    try:
        layout = LayoutReport.objects.get(pk=layout_id, usuario=request.user)
    except LayoutReport.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Layout não encontrado'}, status=404)
    
    try:
        # Se recebeu layout_json atualizado, usar ele
        if layout_json and layout_json.get('bands'):
            layout.layout_json = layout_json
        
        # Debug logs
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"Gerando preview do layout {layout.nome}")
        
        # Gerar preview com dados de exemplo mais realistas
        pdf_bytes = generate_preview_pdf(layout, sample_size=10)
        
        logger.info(f"Preview gerado com sucesso")
        
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="preview_{layout.nome}.pdf"'
        # Adicionar cabeçalhos para melhor compatibilidade com iframes
        response['X-Frame-Options'] = 'SAMEORIGIN'
        response['Content-Security-Policy'] = "frame-ancestors 'self'"
        return response
        
    except Exception as e:
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f"Erro ao gerar preview: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        
        return JsonResponse({
            'success': False,
            'message': f'Erro ao gerar preview: {str(e)}'
        }, status=500)


@login_required
def report_designer_generate(request, pk):
    """Gerar relatório PDF completo com dados reais de uma rota usando busca completa do Firebird"""
    layout = get_object_or_404(LayoutReport, pk=pk, usuario=request.user)
    
    rota_id = request.GET.get('rota_id')
    
    # Buscar dados reais usando nova função que busca completo do Firebird
    dados = []
    rota = None
    
    if rota_id:
        try:
            from rotas.report_renderer import get_full_firebird_data_for_rota
            rota = Rota.objects.get(pk=rota_id, usuario=request.user)
            
            print(f"DEBUG: Gerando relatório para rota {rota.id} com busca completa do Firebird")
            dados = get_full_firebird_data_for_rota(rota)
            
        except Rota.DoesNotExist:
            pass
    
    if not dados:
        # Sem rota específica, buscar últimos pedidos como exemplo
        pedidos = PedidoFirebird.objects.filter(usuario=request.user).order_by('-data_importacao')[:20]
        for pedido in pedidos:
            dados.append(pedido.dados_json or {})
    
    try:
        start_time = time.time()
        
        # Debug logs
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"Gerando relatório {layout.nome} com {len(dados)} registros")
        if dados:
            logger.info(f"Exemplo de dados: {list(dados[0].keys())[:10]}...")
        
        pdf_bytes = generate_pdf(layout, dados)
        elapsed = time.time() - start_time
        
        logger.info(f"Relatório gerado em {elapsed:.2f}s")
        
        # Registrar no histórico
        ReportHistory.objects.create(
            layout=layout,
            usuario=request.user,
            rota=rota,
            total_registros=len(dados),
            total_paginas=0,  # TODO: extrair do PDF
            tempo_geracao=elapsed,
            formato='pdf'
        )
        
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="relatorio_{layout.nome}.pdf"'
        return response
        
    except Exception as e:
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f"Erro ao gerar relatório: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        
        return JsonResponse({
            'success': False,
            'message': f'Erro ao gerar relatório: {str(e)}'
        }, status=500)


@login_required
def report_designer_export(request, pk):
    """Exportar layout de relatório como JSON"""
    layout = get_object_or_404(LayoutReport, pk=pk, usuario=request.user)
    
    export_data = {
        'nome': layout.nome,
        'descricao': layout.descricao,
        'page_size': layout.page_size,
        'orientation': layout.orientation,
        'margins': {
            'top': layout.margin_top,
            'bottom': layout.margin_bottom,
            'left': layout.margin_left,
            'right': layout.margin_right
        },
        'default_font': {
            'family': layout.default_font_family,
            'size': layout.default_font_size
        },
        'layout': layout.layout_json,
        'exported_at': datetime.now().isoformat(),
        'version': layout.versao
    }
    
    response = HttpResponse(
        json.dumps(export_data, ensure_ascii=False, indent=2),
        content_type='application/json'
    )
    response['Content-Disposition'] = f'attachment; filename="layout_{layout.nome}_{layout.id}.json"'
    
    return response


@login_required
@require_POST
def report_designer_export_format(request):
    """Exportar relatório em Excel, CSV ou HTML a partir do layout JSON enviado"""
    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'success': False, 'message': 'JSON inválido'}, status=400)

    layout_id = data.get('layout_id')
    layout_json = data.get('layout_json', {})
    fmt = data.get('format', 'csv').lower()

    if fmt not in ('excel', 'csv', 'html'):
        return JsonResponse({'success': False, 'message': f'Formato "{fmt}" não suportado'}, status=400)

    # Obter layout do banco para query associada
    query_obj = None
    dataset = []
    columns = []
    layout_name = 'relatorio'

    if layout_id:
        try:
            layout_obj = LayoutReport.objects.get(pk=layout_id, usuario=request.user)
            layout_name = layout_obj.nome
            query_obj = layout_obj.query
        except LayoutReport.DoesNotExist:
            pass

    # Executar query para obter dados
    if query_obj:
        try:
            from rotas.models import ConexaoFirebird
            conn_obj = ConexaoFirebird.objects.filter(usuario=request.user, ativo=True).first()
            if conn_obj:
                import fdb
                conn = fdb.connect(
                    host=conn_obj.host, database=conn_obj.database,
                    user=conn_obj.usuario, password=conn_obj.senha,
                    charset='UTF8'
                )
                cur = conn.cursor()
                cur.execute(query_obj.sql_query)
                columns = [d[0] for d in cur.description] if cur.description else []
                rows = cur.fetchall()
                dataset = [dict(zip(columns, row)) for row in rows]
                conn.close()
        except Exception as e:
            # If DB unavailable, export empty with column headers from layout
            pass

    # Derive columns from layout if not from query
    if not columns and layout_json.get('bands'):
        seen = set()
        for band in layout_json['bands']:
            for el in band.get('elements', []):
                field = el.get('data_field') or el.get('aggregate_field')
                if field and field not in seen:
                    seen.add(field)
                    columns.append(field)

    if fmt == 'csv':
        import csv
        import io
        buf = io.StringIO()
        writer = csv.writer(buf)
        if columns:
            writer.writerow(columns)
        for row in dataset:
            writer.writerow([row.get(c, '') for c in columns])
        response = HttpResponse(buf.getvalue(), content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{layout_name}.csv"'
        response.write('\ufeff'.encode('utf-8'))  # BOM for Excel compatibility
        return response

    if fmt == 'html':
        th_cells = ''.join(f'<th>{c}</th>' for c in columns)
        tr_rows = ''
        for i, row in enumerate(dataset):
            bg = '#f9f9f9' if i % 2 else '#ffffff'
            tds = ''.join(f'<td>{row.get(c, "")}</td>' for c in columns)
            tr_rows += f'<tr style="background:{bg}">{tds}</tr>'
        html = f'''<!DOCTYPE html><html><head><meta charset="UTF-8">
<title>{layout_name}</title>
<style>
  body{{font-family:Arial,sans-serif;font-size:12px;margin:20px}}
  h1{{font-size:18px;margin-bottom:12px}}
  table{{border-collapse:collapse;width:100%}}
  th{{background:#0e639c;color:#fff;padding:6px 10px;text-align:left;border:1px solid #999}}
  td{{padding:5px 10px;border:1px solid #ddd}}
</style></head><body>
<h1>{layout_name}</h1>
<table><thead><tr>{th_cells}</tr></thead><tbody>{tr_rows}</tbody></table>
</body></html>'''
        response = HttpResponse(html, content_type='text/html; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{layout_name}.html"'
        return response

    if fmt == 'excel':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            import io as _io

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = layout_name[:31]

            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill('solid', fgColor='0E639C')

            for col_idx, col in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

            for row_idx, row in enumerate(dataset, start=2):
                for col_idx, col in enumerate(columns, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=row.get(col, ''))

            # Auto-width
            for col in ws.columns:
                max_len = max((len(str(c.value)) if c.value else 0) for c in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

            buf = _io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            response = HttpResponse(
                buf.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{layout_name}.xlsx"'
            return response
        except ImportError:
            return JsonResponse({'success': False, 'message': 'openpyxl não instalado. Execute: pip install openpyxl'}, status=500)

    return JsonResponse({'success': False, 'message': 'Formato não implementado'}, status=400)


@login_required
@require_POST
def report_designer_import(request):
    """Importar layout de relatório de JSON"""
    if 'file' not in request.FILES:
        messages.error(request, 'Nenhum arquivo foi enviado')
        return redirect('rotas:report_designer_list')
    
    file = request.FILES['file']
    
    try:
        import_data = json.loads(file.read().decode('utf-8'))
    except (json.JSONDecodeError, UnicodeDecodeError):
        messages.error(request, 'Arquivo JSON inválido')
        return redirect('rotas:report_designer_list')
    
    # Validar estrutura
    required_fields = ['nome', 'layout']
    if not all(field in import_data for field in required_fields):
        messages.error(request, 'Arquivo JSON com estrutura inválida')
        return redirect('rotas:report_designer_list')
    
    # Criar novo layout com dados importados
    margins = import_data.get('margins', {})
    default_font = import_data.get('default_font', {})
    
    layout = LayoutReport.objects.create(
        usuario=request.user,
        nome=f"{import_data['nome']} (Importado)",
        descricao=import_data.get('descricao', ''),
        page_size=import_data.get('page_size', 'A4'),
        orientation=import_data.get('orientation', 'portrait'),
        margin_top=margins.get('top', 20.0),
        margin_bottom=margins.get('bottom', 20.0),
        margin_left=margins.get('left', 15.0),
        margin_right=margins.get('right', 15.0),
        default_font_family=default_font.get('family', 'Helvetica'),
        default_font_size=default_font.get('size', 10),
        layout_json=import_data['layout']
    )
    
    messages.success(request, f'Layout "{layout.nome}" importado com sucesso!')
    return redirect('rotas:report_designer_editor', pk=layout.pk)


# ===== INTEGRAÇÃO COM ROTAS =====

@login_required
def rota_selecionar_layout(request, rota_id):
    """Interface para selecionar layout de relatório para imprimir uma rota"""
    rota = get_object_or_404(Rota, pk=rota_id, usuario=request.user)
    
    # Buscar layouts disponíveis do usuário
    layouts = LayoutReport.objects.filter(
        usuario=request.user, 
        ativo=True
    ).order_by('-data_modificacao')
    
    context = {
        'rota': rota,
        'layouts': layouts,
        'total_layouts': layouts.count()
    }
    
    return render(request, 'rotas/rota_selecionar_layout.html', context)


@login_required 
def rota_gerar_pdf_com_layout(request, rota_id, layout_id):
    """Gerar PDF da rota usando layout personalizado com busca completa do Firebird"""
    rota = get_object_or_404(Rota, pk=rota_id, usuario=request.user)
    layout = get_object_or_404(LayoutReport, pk=layout_id, usuario=request.user, ativo=True)
    
    # Usar nova função que busca dados completos do Firebird
    from rotas.report_renderer import get_full_firebird_data_for_rota
    
    print(f"DEBUG: Gerando PDF da rota {rota.id} com layout {layout.id} usando busca completa do Firebird")
    dados = get_full_firebird_data_for_rota(rota)
    
    try:
        start_time = time.time()
        pdf_bytes = generate_pdf(layout, dados)
        elapsed = time.time() - start_time
        
        # Registrar no histórico
        ReportHistory.objects.create(
            layout=layout,
            usuario=request.user,
            rota=rota,
            total_registros=len(dados),
            total_paginas=0,  # TODO: calcular páginas do PDF
            tempo_geracao=elapsed,
            formato='pdf'
        )
        
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        filename = f"rota_{rota.nome}_{layout.nome}.pdf"
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        return response
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        messages.error(request, f'Erro ao gerar PDF: {str(e)}')
        return redirect('rotas:rota_selecionar_layout', rota_id=rota_id)


@login_required
def rota_preview_layout(request, rota_id, layout_id):
    """Preview rápido de como ficará o relatório da rota com o layout usando busca completa do Firebird"""
    rota = get_object_or_404(Rota, pk=rota_id, usuario=request.user)
    layout = get_object_or_404(LayoutReport, pk=layout_id, usuario=request.user, ativo=True)
    
    # Usar nova função que busca dados completos do Firebird (limitado a 3 para preview)
    from rotas.report_renderer import get_full_firebird_data_for_rota
    
    dados = get_full_firebird_data_for_rota(rota, limit=3)
    
    try:
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"Gerando preview da rota {rota.id} com layout {layout.id}")
        
        pdf_bytes = generate_pdf(layout, dados)
        
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="preview_rota_{rota.id}.pdf"'
        return response
        
    except Exception as e:
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f"Erro ao gerar preview da rota: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        
        return JsonResponse({
            'success': False,
            'message': f'Erro ao gerar preview: {str(e)}'
        }, status=500)
